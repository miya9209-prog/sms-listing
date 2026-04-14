import io
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="광고발송 고객리스트 생성", layout="wide")

TITLE = "광고발송 고객리스트 생성"
TARGET_COLUMNS = {
    "실결제금액": "실결제금액",
    "실결제율": "실결제율",
    "주문당단가": "1회주문당금액",
    "최종주문일": "최종주문일",
    "최종접속일": "최종접속일",
}

NUMERIC_COLUMNS = ["실결제금액", "실결제율", "1회주문당금액"]
DATE_COLUMNS = ["최종주문일", "최종접속일"]
DEFAULT_EXPORT_COLUMNS = [
    "아이디", "이름", "회원등급", "총구매금액", "실결제금액", "실결제율", "1회주문당금액",
    "총 방문횟수(1년 내)", "총 실주문건수", "최종접속일", "최종주문일", "휴대폰번호", "이메일",
    "모바일 메시지 수신여부", "e메일 수신여부", "불량회원"
]


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]

    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    for col in DATE_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
            df[f"{col}_날짜"] = df[col].dt.date

    return df


@st.cache_data(show_spinner=False)
def load_excel(file_bytes: bytes) -> pd.DataFrame:
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    preferred_sheet = "원본" if "원본" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=preferred_sheet)
    return normalize_dataframe(df)



def validate_columns(df: pd.DataFrame):
    missing = [real_name for real_name in TARGET_COLUMNS.values() if real_name not in df.columns]
    return missing



def apply_filters(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    filtered = df.copy()

    amount_min = filters.get("실결제금액_min")
    if amount_min is not None:
        filtered = filtered[filtered["실결제금액"] >= amount_min]

    rate_min = filters.get("실결제율_min")
    if rate_min is not None:
        filtered = filtered[filtered["실결제율"] >= rate_min]

    avg_order_min = filters.get("1회주문당금액_min")
    if avg_order_min is not None:
        filtered = filtered[filtered["1회주문당금액"] >= avg_order_min]

    last_order_until = filters.get("최종주문일_until")
    if last_order_until is not None:
        filtered = filtered[filtered["최종주문일_날짜"].notna() & (filtered["최종주문일_날짜"] <= last_order_until)]

    last_visit_until = filters.get("최종접속일_until")
    if last_visit_until is not None:
        filtered = filtered[filtered["최종접속일_날짜"].notna() & (filtered["최종접속일_날짜"] <= last_visit_until)]

    return filtered



def format_preview_df(df: pd.DataFrame) -> pd.DataFrame:
    preview = df.copy()
    if "실결제율" in preview.columns:
        preview["실결제율"] = preview["실결제율"].map(lambda x: f"{x:.1%}" if pd.notna(x) else "")
    if "실결제금액" in preview.columns:
        preview["실결제금액"] = preview["실결제금액"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "")
    if "1회주문당금액" in preview.columns:
        preview["1회주문당금액"] = preview["1회주문당금액"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "")
    for col in DATE_COLUMNS:
        if col in preview.columns:
            preview[col] = pd.to_datetime(preview[col], errors="coerce").dt.strftime("%Y-%m-%d")
            preview[col] = preview[col].fillna("")
    return preview



def dataframe_to_excel_bytes(df: pd.DataFrame, sort_label: str, sort_column: str, ascending: bool, applied_filters_text: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "광고발송리스트"

    visible_columns = [col for col in DEFAULT_EXPORT_COLUMNS if col in df.columns]
    extra_columns = [col for col in df.columns if col not in visible_columns and not col.endswith("_날짜")]
    ordered_columns = visible_columns + extra_columns

    ws["A1"] = TITLE
    ws["A2"] = f"정렬 기준: {sort_label} ({'오름차순' if ascending else '내림차순'})"
    ws["A3"] = f"적용 조건: {applied_filters_text}"
    ws["A4"] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    title_fill = PatternFill("solid", fgColor="111111")
    sub_fill = PatternFill("solid", fgColor="F3F4F6")
    header_fill = PatternFill("solid", fgColor="E7EEF8")
    thin = Side(style="thin", color="D1D5DB")

    for cell in [ws["A1"], ws["A2"], ws["A3"], ws["A4"]]:
        cell.alignment = Alignment(vertical="center")

    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(ordered_columns), 4))

    for row_idx in [2, 3, 4]:
        ws.cell(row=row_idx, column=1).font = Font(size=10, bold=(row_idx == 2))
        ws.cell(row=row_idx, column=1).fill = sub_fill
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max(len(ordered_columns), 4))

    header_row = 6
    for col_idx, col_name in enumerate(ordered_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    export_df = df[ordered_columns].copy()
    for col in DATE_COLUMNS:
        if col in export_df.columns:
            export_df[col] = pd.to_datetime(export_df[col], errors="coerce").dt.date

    for row_idx, row in enumerate(export_df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            column_name = ordered_columns[col_idx - 1]
            if column_name in ["실결제금액", "총구매금액", "1회주문당금액", "총 사용 적립금", "총적립금", "사용가능 적립금"]:
                cell.number_format = '#,##0'
            elif column_name == "실결제율":
                cell.number_format = '0.0%'
            elif column_name in DATE_COLUMNS or column_name in ["회원 가입일", "생년월일"]:
                cell.number_format = 'yyyy-mm-dd'
            cell.alignment = Alignment(vertical="center")

    if len(export_df) > 0 and len(ordered_columns) > 0:
        ref = f"A{header_row}:{get_column_letter(len(ordered_columns))}{header_row + len(export_df)}"
        table = Table(displayName="AdCustomerList", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

    for idx, column_name in enumerate(ordered_columns, start=1):
        max_len = len(str(column_name))
        for row in range(header_row + 1, min(header_row + len(export_df) + 1, header_row + 300)):
            value = ws.cell(row=row, column=idx).value
            if value is None:
                continue
            display = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)
            max_len = max(max_len, len(display))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 24)

    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


st.title(TITLE)
st.caption("견본 xlsx 파일을 업로드한 뒤, 조건별로 고객을 추려 정렬하고 xlsx 파일로 다시 다운로드할 수 있습니다.")

uploaded_file = st.file_uploader("엑셀 파일 업로드", type=["xlsx"])

with st.expander("지원 조건 보기", expanded=True):
    st.write("- 실결제금액 이상")
    st.write("- 실결제율 이상")
    st.write("- 주문당단가 이상")
    st.write("- 최종주문일: 지정 날짜까지")
    st.write("- 최종접속일: 지정 날짜까지")

if uploaded_file is not None:
    try:
        file_bytes = uploaded_file.getvalue()
        df = load_excel(file_bytes)
        missing = validate_columns(df)

        if missing:
            st.error("필수 컬럼이 부족합니다: " + ", ".join(missing))
            st.stop()

        c1, c2, c3 = st.columns(3)
        with c1:
            amount_min_input = st.number_input("실결제금액 이상", min_value=0, value=0, step=10000, format="%d")
        with c2:
            rate_min_percent = st.number_input("실결제율 이상 (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0)
        with c3:
            avg_order_min_input = st.number_input("주문당단가 이상", min_value=0, value=0, step=1000, format="%d")

        d1, d2 = st.columns(2)
        with d1:
            use_last_order = st.checkbox("최종주문일 조건 사용", value=False)
            last_order_until = st.date_input("최종주문일 지정일까지", value=date.today(), disabled=not use_last_order)
        with d2:
            use_last_visit = st.checkbox("최종접속일 조건 사용", value=False)
            last_visit_until = st.date_input("최종접속일 지정일까지", value=date.today(), disabled=not use_last_visit)

        s1, s2 = st.columns([2, 1])
        with s1:
            sort_label = st.selectbox("정렬 기준", list(TARGET_COLUMNS.keys()), index=0)
        with s2:
            ascending = st.selectbox("정렬 방향", ["내림차순", "오름차순"], index=0) == "오름차순"

        filters = {
            "실결제금액_min": amount_min_input if amount_min_input > 0 else None,
            "실결제율_min": (rate_min_percent / 100.0) if rate_min_percent > 0 else None,
            "1회주문당금액_min": avg_order_min_input if avg_order_min_input > 0 else None,
            "최종주문일_until": last_order_until if use_last_order else None,
            "최종접속일_until": last_visit_until if use_last_visit else None,
        }

        filtered_df = apply_filters(df, filters)
        sort_column = TARGET_COLUMNS[sort_label]
        filtered_df = filtered_df.sort_values(by=sort_column, ascending=ascending, na_position="last").reset_index(drop=True)

        applied_filters = []
        if filters["실결제금액_min"] is not None:
            applied_filters.append(f"실결제금액 ≥ {filters['실결제금액_min']:,}")
        if filters["실결제율_min"] is not None:
            applied_filters.append(f"실결제율 ≥ {filters['실결제율_min']:.1%}")
        if filters["1회주문당금액_min"] is not None:
            applied_filters.append(f"주문당단가 ≥ {filters['1회주문당금액_min']:,}")
        if filters["최종주문일_until"] is not None:
            applied_filters.append(f"최종주문일 ≤ {filters['최종주문일_until'].strftime('%Y-%m-%d')}")
        if filters["최종접속일_until"] is not None:
            applied_filters.append(f"최종접속일 ≤ {filters['최종접속일_until'].strftime('%Y-%m-%d')}")
        applied_filters_text = ", ".join(applied_filters) if applied_filters else "조건 없음"

        m1, m2, m3 = st.columns(3)
        m1.metric("원본 고객 수", f"{len(df):,}명")
        m2.metric("필터 결과", f"{len(filtered_df):,}명")
        m3.metric("정렬 기준", sort_label)

        preview_columns = [col for col in DEFAULT_EXPORT_COLUMNS if col in filtered_df.columns]
        preview_df = format_preview_df(filtered_df[preview_columns].head(200))
        st.dataframe(preview_df, use_container_width=True, height=520)

        excel_bytes = dataframe_to_excel_bytes(filtered_df, sort_label, sort_column, ascending, applied_filters_text)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ad_customer_list_{timestamp}.xlsx"

        st.download_button(
            "엑셀 다운로드 (.xlsx)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception as e:
        st.error(f"파일 처리 중 오류가 발생했습니다: {e}")
else:
    st.info("먼저 견본 또는 실제 고객 xlsx 파일을 업로드해 주세요.")
